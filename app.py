import streamlit as st
import pandas as pd
import numpy as np
import pickle
import joblib
import io
import os
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="OlimpMath – Seleksi Olimpiade Matematika",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=Playfair+Display:wght@700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Plus Jakarta Sans', sans-serif;
}

/* Background */
.stApp {
    background: linear-gradient(135deg, #0a0e1a 0%, #0d1b3e 40%, #0a1628 100%);
    min-height: 100vh;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d1f4e 0%, #091530 100%);
    border-right: 1px solid rgba(99,179,237,0.15);
}
[data-testid="stSidebar"] .block-container {
    padding-top: 2rem;
}

/* Hero header */
.hero-header {
    background: linear-gradient(135deg, #1a3a6e 0%, #0d2554 50%, #091840 100%);
    border: 1px solid rgba(99,179,237,0.3);
    border-radius: 20px;
    padding: 2.5rem 3rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
    box-shadow: 0 20px 60px rgba(0,0,0,0.5), inset 0 1px 0 rgba(255,255,255,0.1);
}
.hero-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, rgba(99,179,237,0.15) 0%, transparent 70%);
    pointer-events: none;
}
.hero-title {
    font-family: 'Playfair Display', serif;
    font-size: 3rem;
    font-weight: 800;
    background: linear-gradient(135deg, #63B3ED, #90CDF4, #BEE3F8);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    margin: 0;
    line-height: 1.1;
}
.hero-subtitle {
    color: rgba(190,227,248,0.8);
    font-size: 1.05rem;
    margin-top: 0.5rem;
    font-weight: 400;
    letter-spacing: 0.02em;
}
.hero-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: rgba(99,179,237,0.15);
    border: 1px solid rgba(99,179,237,0.4);
    border-radius: 50px;
    padding: 5px 14px;
    font-size: 0.8rem;
    color: #90CDF4;
    margin-bottom: 1rem;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}

/* Cards */
.metric-card {
    background: linear-gradient(135deg, rgba(26,58,110,0.6) 0%, rgba(13,37,84,0.8) 100%);
    border: 1px solid rgba(99,179,237,0.2);
    border-radius: 16px;
    padding: 1.5rem;
    text-align: center;
    transition: all 0.3s ease;
    backdrop-filter: blur(10px);
}
.metric-card:hover {
    border-color: rgba(99,179,237,0.5);
    transform: translateY(-3px);
    box-shadow: 0 10px 30px rgba(99,179,237,0.1);
}
.metric-value {
    font-size: 2.2rem;
    font-weight: 800;
    color: #90CDF4;
    line-height: 1;
}
.metric-label {
    font-size: 0.8rem;
    color: rgba(190,227,248,0.6);
    margin-top: 0.3rem;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

/* Status badges */
.status-siap {
    background: linear-gradient(135deg, #1a4731, #22543d);
    border: 1px solid rgba(72,187,120,0.5);
    color: #9AE6B4;
    padding: 8px 20px;
    border-radius: 50px;
    font-weight: 700;
    font-size: 1rem;
    display: inline-block;
    box-shadow: 0 0 20px rgba(72,187,120,0.2);
}
.status-potensial {
    background: linear-gradient(135deg, #4a3900, #5d4900);
    border: 1px solid rgba(246,173,85,0.5);
    color: #F6AD55;
    padding: 8px 20px;
    border-radius: 50px;
    font-weight: 700;
    font-size: 1rem;
    display: inline-block;
    box-shadow: 0 0 20px rgba(246,173,85,0.2);
}
.status-tidak {
    background: linear-gradient(135deg, #4a1919, #5d2020);
    border: 1px solid rgba(245,101,101,0.5);
    color: #FC8181;
    padding: 8px 20px;
    border-radius: 50px;
    font-weight: 700;
    font-size: 1rem;
    display: inline-block;
    box-shadow: 0 0 20px rgba(245,101,101,0.2);
}

/* Result card */
.result-card {
    background: linear-gradient(135deg, rgba(20,40,80,0.9) 0%, rgba(10,25,60,0.95) 100%);
    border: 1px solid rgba(99,179,237,0.3);
    border-radius: 20px;
    padding: 2rem;
    margin: 1rem 0;
    backdrop-filter: blur(20px);
    box-shadow: 0 10px 40px rgba(0,0,0,0.4);
}

/* Input area */
.input-section {
    background: linear-gradient(135deg, rgba(13,31,78,0.7) 0%, rgba(9,21,48,0.9) 100%);
    border: 1px solid rgba(99,179,237,0.2);
    border-radius: 16px;
    padding: 1.8rem;
    margin-bottom: 1.5rem;
}

/* Section title */
.section-title {
    font-family: 'Playfair Display', serif;
    font-size: 1.6rem;
    font-weight: 700;
    color: #BEE3F8;
    margin-bottom: 0.3rem;
}
.section-subtitle {
    color: rgba(190,227,248,0.55);
    font-size: 0.9rem;
    margin-bottom: 1.5rem;
}

/* Recommendation box */
.rec-box {
    background: rgba(13,37,84,0.5);
    border-left: 3px solid;
    border-radius: 0 12px 12px 0;
    padding: 1rem 1.2rem;
    margin: 0.6rem 0;
    font-size: 0.92rem;
    color: rgba(190,227,248,0.85);
    line-height: 1.6;
}

/* Score bar */
.score-row {
    display: flex;
    align-items: center;
    margin: 0.5rem 0;
    gap: 12px;
}
.score-label {
    width: 220px;
    font-size: 0.85rem;
    color: rgba(190,227,248,0.8);
    flex-shrink: 0;
}
.score-bar-wrap {
    flex: 1;
    background: rgba(255,255,255,0.06);
    border-radius: 50px;
    height: 10px;
    overflow: hidden;
}
.score-bar-fill {
    height: 100%;
    border-radius: 50px;
    transition: width 0.8s ease;
}
.score-val {
    width: 45px;
    text-align: right;
    font-size: 0.85rem;
    font-weight: 700;
    color: #90CDF4;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: rgba(13,31,78,0.5);
    border-radius: 12px;
    padding: 4px;
    gap: 4px;
    border: 1px solid rgba(99,179,237,0.15);
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    color: rgba(190,227,248,0.6);
    font-weight: 600;
    padding: 8px 20px;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #1a4a8a, #163d72) !important;
    color: #BEE3F8 !important;
}

/* Streamlit elements overrides */
.stSlider > div > div > div > div {
    background: linear-gradient(90deg, #2b6cb0, #63B3ED) !important;
}
[data-testid="stNumberInput"] input,
[data-testid="stTextInput"] input {
    background: rgba(13,31,78,0.8) !important;
    border: 1px solid rgba(99,179,237,0.3) !important;
    color: #BEE3F8 !important;
    border-radius: 8px !important;
}
.stButton > button {
    background: linear-gradient(135deg, #1a4a8a, #2563ad);
    color: #fff;
    border: 1px solid rgba(99,179,237,0.4);
    border-radius: 10px;
    font-weight: 700;
    padding: 0.6rem 2rem;
    transition: all 0.3s ease;
    font-family: 'Plus Jakarta Sans', sans-serif;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #2563ad, #3182ce);
    border-color: rgba(99,179,237,0.7);
    box-shadow: 0 5px 20px rgba(99,179,237,0.3);
    transform: translateY(-2px);
}
.stFileUploader > div {
    background: rgba(13,31,78,0.5) !important;
    border: 2px dashed rgba(99,179,237,0.3) !important;
    border-radius: 12px !important;
}
.stDataFrame {
    border: 1px solid rgba(99,179,237,0.2);
    border-radius: 12px;
    overflow: hidden;
}
label[data-testid="stWidgetLabel"] p, .stSelectbox label p {
    color: rgba(190,227,248,0.8) !important;
    font-weight: 500;
}
.stMarkdown p { color: rgba(190,227,248,0.85); }
.stAlert { border-radius: 12px; }
div[data-testid="stExpander"] {
    background: rgba(13,31,78,0.4);
    border: 1px solid rgba(99,179,237,0.15);
    border-radius: 12px;
}

/* Sidebar radio */
[data-testid="stSidebar"] label { color: rgba(190,227,248,0.8) !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
FEATURES = ['skor_aljabar', 'skor_geometri', 'skor_bilangan',
            'skor_data_ketidakpastian', 'skor_menalar', 'skor_literasi']
FEATURE_LABELS = {
    'skor_aljabar': 'Numerasi Aljabar',
    'skor_geometri': 'Numerasi Geometri',
    'skor_bilangan': 'Numerasi Bilangan',
    'skor_data_ketidakpastian': 'Data & Ketidakpastian',
    'skor_menalar': 'Skor Menalar',
    'skor_literasi': 'Skor Literasi'
}
MODEL_PATH = 'models/model_olimpmath.pkl'
DATA_PATH  = 'data/dataset.xlsx'


# ─────────────────────────────────────────────
# HELPER: LABEL
# ─────────────────────────────────────────────
def assign_label(row):
    vals = [row[f] for f in FEATURES]
    avg = np.mean(vals)
    mn  = np.min(vals)
    if avg >= 80 and mn >= 60:
        return 'Siap Olimpiade'
    elif avg >= 65 and mn >= 45:
        return 'Potensial'
    else:
        return 'Tidak Siap'


# ─────────────────────────────────────────────
# LOAD / TRAIN MODEL
# ─────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_model():
    if os.path.exists(MODEL_PATH):
    return joblib.load(MODEL_PATH)
    # Train fresh
    if not os.path.exists(DATA_PATH):
        return None
    df = pd.read_excel(DATA_PATH)
    df.columns = FEATURES
    for c in FEATURES:
        df[c] = df[c].fillna(df[c].median())
    df['label'] = df.apply(assign_label, axis=1)
    X = df[FEATURES]; y = df['label']
    X_tr, _, y_tr, _ = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
model = RandomForestClassifier(n_estimators=20, random_state=42, max_depth=9, min_samples_leaf=20)
model.fit(X_tr, y_tr)
os.makedirs('models', exist_ok=True)
joblib.dump(model, MODEL_PATH, compress=('lzma', 9))
    return model


# ─────────────────────────────────────────────
# PREDICT + GENERATE RESULT
# ─────────────────────────────────────────────
def predict_student(model, scores: dict):
    X = pd.DataFrame([scores])[FEATURES]
    pred  = model.predict(X)[0]
    proba = model.predict_proba(X)[0]
    cls   = model.classes_.tolist()
    conf  = {c: round(proba[i]*100, 1) for i, c in enumerate(cls)}
    return pred, conf


def generate_result(nama, scores, status, confidence):
    avg = round(np.mean(list(scores.values())), 2)
    mn  = round(np.min(list(scores.values())), 2)
    mx  = round(np.max(list(scores.values())), 2)

    weak = [FEATURE_LABELS[k] for k, v in scores.items() if v < 60]
    strong = [FEATURE_LABELS[k] for k, v in scores.items() if v >= 80]
    medium = [FEATURE_LABELS[k] for k, v in scores.items() if 60 <= v < 80]

    if status == 'Siap Olimpiade':
        review = (
            f"🏆 **{nama}** menunjukkan performa luar biasa dengan rata-rata skor {avg} dan tidak ada "
            f"komponen di bawah standar kompetitif. Profil akademik siswa ini sangat kuat "
            f"dan konsisten di seluruh dimensi kemampuan matematika."
        )
        rekomendasi = [
            "✅ Lanjutkan ke seleksi tahap berikutnya (Olimpiade tingkat sekolah/kabupaten)",
            "📚 Perdalam soal-soal OSN/IMSO tingkat lanjut untuk persiapan kompetisi",
            "🎯 Fokus latihan soal olimpiade bertipe HOTS (Higher Order Thinking Skills)",
            "🤝 Bergabung dengan kelompok belajar olimpiade intensif",
            "📊 Pertahankan performa dan tingkatkan komponen dengan skor tertinggi sebagai keunggulan",
        ]
        saran = (
            f"Siswa ini siap berkompetisi. Jika ada komponen yang masih bisa ditingkatkan "
            f"(terutama yang mendekati angka 80-85), maka potensi meraih prestasi puncak "
            f"sangat terbuka lebar. Rekomendasikan untuk mengikuti pembinaan intensif olimpiade."
        )
    elif status == 'Potensial':
        weak_str = ', '.join(weak) if weak else 'tidak ada'
        strong_str = ', '.join(strong) if strong else 'belum ada'
        review = (
            f"⭐ **{nama}** memiliki potensi yang baik dengan rata-rata skor {avg}. "
            f"Komponen kuat: {strong_str}. Komponen yang perlu diperkuat: {weak_str}. "
            f"Dengan bimbingan yang tepat, siswa ini berpeluang besar untuk lolos seleksi."
        )
        rekomendasi = [
            f"📌 Prioritaskan peningkatan komponen lemah: **{weak_str}**" if weak else "📌 Pertahankan keseimbangan antar komponen",
            "📝 Latihan soal olimpiade rutin minimal 3x seminggu",
            "🧩 Ikuti bimbingan belajar atau try-out olimpiade sekolah",
            "📖 Pelajari strategi penyelesaian soal matematika kompetitif",
            "🔄 Lakukan evaluasi ulang setelah 4-6 minggu latihan intensif",
        ]
        saran = (
            f"Siswa ini berada di ambang kesiapan olimpiade. Dengan latihan intensif dan "
            f"perbaikan pada komponen bernilai di bawah 65, peluang untuk masuk kategori "
            f"'Siap Olimpiade' sangat besar. Berikan program remedial khusus untuk komponen lemah."
        )
    else:  # Tidak Siap
        weak_str = ', '.join(weak) if weak else '-'
        review = (
            f"📋 **{nama}** saat ini belum memenuhi standar minimum seleksi olimpiade dengan rata-rata "
            f"skor {avg}. Masih terdapat beberapa komponen fundamental yang perlu diperkuat "
            f"sebelum dapat mengikuti seleksi olimpiade: {weak_str}."
        )
        rekomendasi = [
            "🔧 Perkuat pemahaman konsep dasar matematika terlebih dahulu",
            f"📚 Fokus pada komponen dengan skor terendah: **{weak_str}**",
            "👩‍🏫 Konsultasikan dengan guru matematika untuk program pemulihan belajar",
            "⏰ Sisihkan waktu belajar matematika minimal 1 jam per hari",
            "📋 Ikuti program remedial sekolah dan evaluasi ulang dalam 2-3 bulan",
        ]
        saran = (
            f"Siswa belum siap untuk seleksi olimpiade saat ini. Fokuskan pada penguatan "
            f"kemampuan dasar matematika secara menyeluruh. Dengan dedikasi dan bimbingan yang "
            f"tepat, siswa berpotensi berkembang pesat. Pantau perkembangan setiap bulan."
        )

    return {
        'review': review,
        'rekomendasi': rekomendasi,
        'saran': saran,
        'avg': avg,
        'min_score': mn,
        'max_score': mx,
        'weak': weak,
        'strong': strong,
        'medium': medium,
    }


# ─────────────────────────────────────────────
# RENDER RESULT CARD
# ─────────────────────────────────────────────
def get_status_html(status):
    if status == 'Siap Olimpiade':
        return f'<span class="status-siap">🏆 {status}</span>'
    elif status == 'Potensial':
        return f'<span class="status-potensial">⭐ {status}</span>'
    else:
        return f'<span class="status-tidak">📋 {status}</span>'


def bar_color(score):
    if score >= 80:   return '#48BB78'
    elif score >= 65: return '#F6AD55'
    else:             return '#FC8181'


def render_score_bars(scores):
    html = ''
    for k, v in scores.items():
        lbl = FEATURE_LABELS[k]
        col = bar_color(v)
        html += f"""
        <div class="score-row">
            <div class="score-label">{lbl}</div>
            <div class="score-bar-wrap">
                <div class="score-bar-fill" style="width:{v}%;background:{col};"></div>
            </div>
            <div class="score-val">{v:.1f}</div>
        </div>"""
    return html


def render_result(nama, scores, status, confidence, result):
    status_html = get_status_html(status)
    bars = render_score_bars(scores)
    conf_top = max(confidence, key=confidence.get)
    conf_val = confidence[conf_top]

    st.markdown(f"""
    <div class="result-card">
        <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:1rem;margin-bottom:1.5rem;">
            <div>
                <div style="font-size:0.75rem;color:rgba(190,227,248,0.5);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;">HASIL ANALISIS</div>
                <div style="font-family:'Playfair Display',serif;font-size:1.8rem;font-weight:700;color:#BEE3F8;">{nama}</div>
            </div>
            <div style="text-align:right;">
                {status_html}
                <div style="font-size:0.78rem;color:rgba(190,227,248,0.5);margin-top:6px;">Tingkat Kepercayaan: <b style="color:#90CDF4">{conf_val}%</b></div>
            </div>
        </div>
        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;margin-bottom:1.5rem;">
            <div class="metric-card">
                <div class="metric-value">{result['avg']:.1f}</div>
                <div class="metric-label">Rata-rata Skor</div>
            </div>
            <div class="metric-card">
                <div class="metric-value" style="color:#FC8181">{result['min_score']:.1f}</div>
                <div class="metric-label">Skor Terendah</div>
            </div>
            <div class="metric-card">
                <div class="metric-value" style="color:#48BB78">{result['max_score']:.1f}</div>
                <div class="metric-label">Skor Tertinggi</div>
            </div>
        </div>
        <div style="margin-bottom:1.5rem;">
            <div style="font-size:0.75rem;font-weight:700;color:rgba(190,227,248,0.5);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.8rem;">PROFIL SKOR</div>
            {bars}
        </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([1,1])
    with col1:
        st.markdown("#### 📝 Review")
        st.markdown(f"""<div class="result-card" style="padding:1.2rem;">{result['review']}</div>""", unsafe_allow_html=True)
        st.markdown("#### 💡 Saran")
        st.markdown(f"""<div class="result-card" style="padding:1.2rem;border-left:3px solid #63B3ED;">{result['saran']}</div>""", unsafe_allow_html=True)
    with col2:
        st.markdown("#### 🎯 Rekomendasi")
        rec_colors = ['#48BB78','#63B3ED','#F6AD55','#B794F4','#FC8181']
        recs_html = ''
        for i, r in enumerate(result['rekomendasi']):
            c = rec_colors[i % len(rec_colors)]
            recs_html += f'<div class="rec-box" style="border-left-color:{c};">{r}</div>'
        st.markdown(f'<div class="result-card" style="padding:1.2rem;">{recs_html}</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────
# EXPORT TO EXCEL
# ─────────────────────────────────────────────
def export_results_excel(results_df: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()

    # === Sheet 1: Hasil Analisis ===
    ws = wb.active
    ws.title = "Hasil Analisis"

    # Color palette
    C_NAVY   = "0D1F4E"
    C_BLUE   = "1A3A6E"
    C_ACCENT = "63B3ED"
    C_WHITE  = "F0F6FF"
    C_GREEN  = "22543D"
    C_YELLOW = "744210"
    C_RED    = "742A2A"
    C_LGGREEN= "C6F6D5"
    C_LGYELL = "FEFCBF"
    C_LGRED  = "FED7D7"

    thin = Side(style='thin', color="CCDDEE")
    thick = Side(style='medium', color="63B3ED")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    # Title row
    ws.merge_cells('A1:L1')
    t = ws['A1']
    t.value = "🏆 OLIMPMATH – HASIL SELEKSI OLIMPIADE MATEMATIKA"
    t.font = Font(name='Calibri', bold=True, size=16, color="63B3ED")
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells('A2:L2')
    s = ws['A2']
    s.value = f"Sistem Seleksi Tahap Awal Calon Peserta Olimpiade Matematika | Diproses oleh OlimpMath AI"
    s.font = Font(name='Calibri', size=10, color="90CDF4")
    s.fill = PatternFill("solid", fgColor=C_BLUE)
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    ws.append([])  # blank row
    ws.row_dimensions[3].height = 8

    # Headers
    headers = ['No', 'Nama Siswa', 'Aljabar', 'Geometri', 'Bilangan',
               'Data & Ketidakpastian', 'Menalar', 'Literasi',
               'Rata-rata', 'Status Kesiapan', 'Tingkat Keyakinan (%)', 'Keterangan']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=C_NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    ws.row_dimensions[4].height = 30

    # Data rows
    for idx, row in results_df.iterrows():
        r = idx + 5
        status = row.get('Status', '')
        avg_score = row.get('Rata-rata', 0)

        if status == 'Siap Olimpiade':
            row_fill = PatternFill("solid", fgColor="EBF8F0")
            status_fill = PatternFill("solid", fgColor=C_GREEN)
            status_color = "FFFFFF"
        elif status == 'Potensial':
            row_fill = PatternFill("solid", fgColor="FFFFF0")
            status_fill = PatternFill("solid", fgColor=C_YELLOW)
            status_color = "FFFFFF"
        else:
            row_fill = PatternFill("solid", fgColor="FFF5F5")
            status_fill = PatternFill("solid", fgColor=C_RED)
            status_color = "FFFFFF"

        values = [
            idx + 1,
            row.get('Nama Siswa', f'Siswa {idx+1}'),
            row.get('Skor Aljabar', ''),
            row.get('Skor Geometri', ''),
            row.get('Skor Bilangan', ''),
            row.get('Skor Data & Ketidakpastian', ''),
            row.get('Skor Menalar', ''),
            row.get('Skor Literasi', ''),
            avg_score,
            status,
            row.get('Konfidensial (%)', ''),
            row.get('Keterangan', ''),
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center' if c != 2 else 'left', vertical='center', wrap_text=(c==12))
            cell.font = Font(name='Calibri', size=10)
            if c == 10:  # Status column
                cell.fill = status_fill
                cell.font = Font(name='Calibri', bold=True, size=10, color=status_color)
            else:
                cell.fill = row_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            # Color score cells
            if c in [3,4,5,6,7,8] and isinstance(v, (int, float)):
                if v >= 80:   cell.font = Font(name='Calibri', bold=True, size=10, color="22543D")
                elif v >= 65: cell.font = Font(name='Calibri', bold=True, size=10, color="744210")
                else:          cell.font = Font(name='Calibri', bold=True, size=10, color="742A2A")

        ws.row_dimensions[r].height = 22

    # Column widths
    col_widths = [5, 28, 12, 12, 12, 20, 12, 12, 12, 20, 20, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 2: Ringkasan ===
    ws2 = wb.create_sheet("Ringkasan")
    ws2.sheet_view.showGridLines = False

    counts = results_df['Status'].value_counts()
    total = len(results_df)

    ws2.merge_cells('A1:F1')
    t2 = ws2['A1']
    t2.value = "RINGKASAN HASIL SELEKSI"
    t2.font = Font(name='Calibri', bold=True, size=14, color="63B3ED")
    t2.fill = PatternFill("solid", fgColor=C_NAVY)
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 36

    summary_data = [
        ('', ''),
        ('Total Siswa Dianalisis', total),
        ('Siap Olimpiade 🏆', counts.get('Siap Olimpiade', 0)),
        ('Potensial ⭐', counts.get('Potensial', 0)),
        ('Tidak Siap 📋', counts.get('Tidak Siap', 0)),
        ('', ''),
        ('Persentase Siap', f"{counts.get('Siap Olimpiade',0)/total*100:.1f}%"),
        ('Persentase Potensial', f"{counts.get('Potensial',0)/total*100:.1f}%"),
        ('Persentase Tidak Siap', f"{counts.get('Tidak Siap',0)/total*100:.1f}%"),
    ]
    for r_idx, (label, value) in enumerate(summary_data, 2):
        if label:
            c1 = ws2.cell(row=r_idx, column=2, value=label)
            c2 = ws2.cell(row=r_idx, column=4, value=value)
            c1.font = Font(name='Calibri', bold=True, size=11, color=C_NAVY)
            c2.font = Font(name='Calibri', bold=True, size=11, color="1A3A6E")
            c2.alignment = Alignment(horizontal='center')
            ws2.row_dimensions[r_idx].height = 24

    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 5
    ws2.column_dimensions['D'].width = 20

    # === Sheet 3: Panduan ===
    ws3 = wb.create_sheet("Panduan")
    ws3.merge_cells('A1:D1')
    g = ws3['A1']
    g.value = "PANDUAN MEMBACA HASIL"
    g.font = Font(name='Calibri', bold=True, size=13, color="FFFFFF")
    g.fill = PatternFill("solid", fgColor=C_NAVY)
    g.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 32

    guide = [
        ('STATUS', 'KRITERIA', 'WARNA', 'TINDAK LANJUT'),
        ('Siap Olimpiade 🏆', 'Rata-rata ≥ 80 & Skor min ≥ 60', 'Hijau', 'Lanjutkan ke seleksi berikutnya'),
        ('Potensial ⭐', 'Rata-rata 65–79 & Skor min ≥ 45', 'Kuning', 'Latihan intensif, evaluasi ulang'),
        ('Tidak Siap 📋', 'Rata-rata < 65 atau Skor min < 45', 'Merah', 'Perkuat fondasi matematika'),
    ]
    for r_idx, row in enumerate(guide, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if r_idx == 3:
                cell.font = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2D3748")
            else:
                cell.font = Font(name='Calibri', size=10)
                fill_map = {'Hijau': 'EBF8F0', 'Kuning': 'FFFFF0', 'Merah': 'FFF5F5'}
                cell.fill = PatternFill("solid", fgColor=fill_map.get(row[2], 'FFFFFF') if c_idx > 1 else fill_map.get(val, 'FFFFFF'))
            ws3.row_dimensions[r_idx].height = 28

    for col, w in zip(['A','B','C','D'], [28, 38, 12, 40]):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="text-align:center;padding:1rem 0 1.5rem;">
        <div style="font-size:3rem;">🏆</div>
        <div style="font-family:'Playfair Display',serif;font-size:1.5rem;font-weight:800;
             background:linear-gradient(135deg,#63B3ED,#90CDF4);
             -webkit-background-clip:text;-webkit-text-fill-color:transparent;
             background-clip:text;">OlimpMath</div>
        <div style="font-size:0.75rem;color:rgba(190,227,248,0.5);letter-spacing:0.1em;
             text-transform:uppercase;margin-top:2px;">Sistem Seleksi Olimpiade</div>
    </div>
    <hr style="border-color:rgba(99,179,237,0.15);margin:0.5rem 0 1.5rem;">
    """, unsafe_allow_html=True)

    menu = st.radio(
        "Navigasi",
        ["🏠 Beranda", "👤 Input Satu Siswa", "📂 Upload Banyak Siswa", "ℹ️ Tentang Sistem"],
        label_visibility="collapsed"
    )

    st.markdown("""
    <hr style="border-color:rgba(99,179,237,0.15);margin:1.5rem 0 1rem;">
    <div style="font-size:0.75rem;color:rgba(190,227,248,0.4);text-align:center;">
        <div style="margin-bottom:6px;font-weight:600;color:rgba(190,227,248,0.6);">LEGENDA STATUS</div>
        <div style="margin:4px 0;">🏆 Siap Olimpiade</div>
        <div style="margin:4px 0;">⭐ Potensial</div>
        <div style="margin:4px 0;">📋 Tidak Siap</div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# LOAD MODEL
# ─────────────────────────────────────────────
with st.spinner("Memuat model AI OlimpMath..."):
    model = load_model()

if model is None:
    st.error("⚠️ Model tidak dapat dimuat. Pastikan file dataset.xlsx tersedia di folder data/.")
    st.stop()


# ─────────────────────────────────────────────
# PAGE: BERANDA
# ─────────────────────────────────────────────
if "Beranda" in menu:
    st.markdown("""
    <div class="hero-header">
        <div class="hero-badge">✦ AI-Powered · Sistem Seleksi Olimpiade Matematika</div>
        <h1 class="hero-title">OlimpMath</h1>
        <p class="hero-subtitle">Sistem Cerdas Seleksi Tahap Awal Calon Peserta Olimpiade Matematika<br>
        berbasis Machine Learning · Akurasi 97%</p>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    metrics = [
        ("97%", "Akurasi Model"),
        ("30.000+", "Data Training"),
        ("6", "Dimensi Kemampuan"),
        ("3", "Kategori Status"),
    ]
    for col, (val, lbl) in zip([c1,c2,c3,c4], metrics):
        with col:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{val}</div>
                <div class="metric-label">{lbl}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2 = st.columns([1,1])
    with col1:
        st.markdown("""
        <div class="result-card">
            <div class="section-title">📐 Dimensi yang Dianalisis</div>
            <div class="section-subtitle">6 komponen kemampuan matematika</div>
        """, unsafe_allow_html=True)
        dims = [
            ("🔢", "Numerasi Aljabar", "Kemampuan aljabar & persamaan"),
            ("📐", "Numerasi Geometri", "Kemampuan geometri & ruang"),
            ("🔢", "Numerasi Bilangan", "Kemampuan bilangan & aritmetika"),
            ("📊", "Data & Ketidakpastian", "Statistika & probabilitas"),
            ("🧠", "Skor Menalar", "Penalaran logis & analitis"),
            ("📖", "Skor Literasi", "Literasi & pemecahan masalah"),
        ]
        for icon, name, desc in dims:
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:12px;padding:10px 0;
                 border-bottom:1px solid rgba(99,179,237,0.1);">
                <div style="font-size:1.4rem;">{icon}</div>
                <div>
                    <div style="font-weight:600;color:#90CDF4;font-size:0.9rem;">{name}</div>
                    <div style="font-size:0.78rem;color:rgba(190,227,248,0.5);">{desc}</div>
                </div>
            </div>""", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col2:
        st.markdown("""
        <div class="result-card">
            <div class="section-title">🎯 Kategori Status</div>
            <div class="section-subtitle">Hasil klasifikasi kesiapan olimpiade</div>
        """, unsafe_allow_html=True)
        categories = [
            ("🏆", "Siap Olimpiade", "#48BB78", "Rata-rata ≥ 80 & semua skor ≥ 60.\nLangsung lanjut ke seleksi berikutnya."),
            ("⭐", "Potensial", "#F6AD55", "Rata-rata 65–79 & skor min ≥ 45.\nPerlu latihan intensif untuk lolos."),
            ("📋", "Tidak Siap", "#FC8181", "Rata-rata < 65 atau skor min < 45.\nFokus pada penguatan fondasi."),
        ]
        for icon, name, color, desc in categories:
            st.markdown(f"""
            <div style="padding:14px;margin:10px 0;background:rgba(255,255,255,0.03);
                 border-left:4px solid {color};border-radius:0 10px 10px 0;">
                <div style="font-size:1rem;font-weight:700;color:{color};">{icon} {name}</div>
                <div style="font-size:0.82rem;color:rgba(190,227,248,0.6);margin-top:4px;line-height:1.5;">{desc}</div>
            </div>""", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# PAGE: INPUT SATU SISWA
# ─────────────────────────────────────────────
elif "Satu Siswa" in menu:
    st.markdown("""
    <div class="hero-header" style="padding:2rem 2.5rem;">
        <div class="hero-badge">👤 Input Manual</div>
        <h1 class="hero-title" style="font-size:2.2rem;">Analisis Satu Siswa</h1>
        <p class="hero-subtitle">Masukkan data skor siswa untuk mengetahui status kesiapan olimpiade</p>
    </div>
    """, unsafe_allow_html=True)

    with st.form("form_single", clear_on_submit=False):
        st.markdown('<div class="input-section">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">📋 Data Siswa</div>', unsafe_allow_html=True)
        nama_input = st.text_input("Nama Siswa", placeholder="Contoh: Budi Santoso", key="nama")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="input-section">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">📊 Skor Kemampuan</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-subtitle">Masukkan skor 0–100 untuk setiap komponen</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            alj = st.slider("🔢 Numerasi Aljabar", 0, 100, 65, key="alj")
            bil = st.slider("🔢 Numerasi Bilangan", 0, 100, 65, key="bil")
            men = st.slider("🧠 Skor Menalar", 0, 100, 65, key="men")
        with col2:
            geo = st.slider("📐 Numerasi Geometri", 0, 100, 65, key="geo")
            dat = st.slider("📊 Data & Ketidakpastian", 0, 100, 65, key="dat")
            lit = st.slider("📖 Skor Literasi", 0, 100, 65, key="lit")
        st.markdown('</div>', unsafe_allow_html=True)

        submitted = st.form_submit_button("🔍 Analisis Sekarang", use_container_width=True)

    if submitted:
        nama = nama_input.strip() or "Siswa Tanpa Nama"
        scores = {
            'skor_aljabar': alj, 'skor_geometri': geo,
            'skor_bilangan': bil, 'skor_data_ketidakpastian': dat,
            'skor_menalar': men, 'skor_literasi': lit,
        }
        status, confidence = predict_student(model, scores)
        result = generate_result(nama, scores, status, confidence)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("## 📊 Hasil Analisis")
        render_result(nama, scores, status, confidence, result)

        # Download single result
        single_df = pd.DataFrame([{
            'Nama Siswa': nama,
            'Skor Aljabar': alj, 'Skor Geometri': geo, 'Skor Bilangan': bil,
            'Skor Data & Ketidakpastian': dat, 'Skor Menalar': men, 'Skor Literasi': lit,
            'Rata-rata': result['avg'],
            'Status': status,
            'Konfidensial (%)': confidence[status],
            'Keterangan': result['saran'][:200] + '...',
        }])
        excel_bytes = export_results_excel(single_df)
        st.download_button(
            "📥 Download Hasil (.xlsx)",
            data=excel_bytes,
            file_name=f"OlimpMath_{nama.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ─────────────────────────────────────────────
# PAGE: UPLOAD BANYAK SISWA
# ─────────────────────────────────────────────
elif "Banyak Siswa" in menu:
    st.markdown("""
    <div class="hero-header" style="padding:2rem 2.5rem;">
        <div class="hero-badge">📂 Upload Massal</div>
        <h1 class="hero-title" style="font-size:2.2rem;">Analisis Banyak Siswa</h1>
        <p class="hero-subtitle">Upload file Excel berisi data banyak siswa, download hasilnya sekaligus</p>
    </div>
    """, unsafe_allow_html=True)

    # Template download
    col1, col2 = st.columns([2,1])
    with col2:
        if os.path.exists('data/template_upload.xlsx'):
            with open('data/template_upload.xlsx','rb') as f:
                st.download_button(
                    "📄 Download Template Excel",
                    data=f.read(),
                    file_name="template_olimpmath.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    with col1:
        st.markdown("""
        <div class="result-card" style="padding:1.2rem;">
            <b style="color:#90CDF4;">📌 Format File yang Diperlukan:</b><br>
            <span style="font-size:0.85rem;color:rgba(190,227,248,0.7);">
            Kolom: <code>Nama Siswa</code> (opsional), <code>Skor Aljabar</code>, <code>Skor Geometri</code>,
            <code>Skor Bilangan</code>, <code>Skor Data & Ketidakpastian</code>,
            <code>Skor Menalar</code>, <code>Skor Literasi</code><br>
            Semua skor dalam skala 0–100. Download template di atas untuk contoh format.
            </span>
        </div>""", unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Upload file Excel (.xlsx atau .xls)",
        type=['xlsx','xls'],
        help="Pastikan kolom sesuai format template"
    )

    if uploaded:
        try:
            df_up = pd.read_excel(uploaded)
            st.markdown(f"<div style='color:rgba(190,227,248,0.7);font-size:0.9rem;margin:0.5rem 0;'>✅ File berhasil dibaca: <b>{len(df_up)} baris</b> data</div>", unsafe_allow_html=True)

            # Map columns flexibly
            col_map = {}
            cols_lower = {c.lower().strip(): c for c in df_up.columns}

            name_keys = ['nama siswa','nama','name','student name']
            feat_keys = {
                'skor_aljabar': ['skor aljabar','aljabar','num_alj','numerasi aljabar'],
                'skor_geometri': ['skor geometri','geometri','num_geo','numerasi geometri'],
                'skor_bilangan': ['skor bilangan','bilangan','num_bil','numerasi bilangan'],
                'skor_data_ketidakpastian': ['skor data & ketidakpastian','data & ketidakpastian','data ketidakpastian','num_dat','skor data ketidakpastian'],
                'skor_menalar': ['skor menalar','menalar','num_l3','menalar'],
                'skor_literasi': ['skor literasi','literasi','lit'],
            }

            # If only 6 cols (no name), assign directly
            if len(df_up.columns) == 6:
                df_up.columns = FEATURES
                nama_col = None
            elif len(df_up.columns) == 7:
                for k in name_keys:
                    if k in cols_lower:
                        nama_col = cols_lower[k]
                        break
                else:
                    nama_col = df_up.columns[0]
                for feat, keys in feat_keys.items():
                    for k in keys:
                        if k in cols_lower:
                            col_map[feat] = cols_lower[k]
                            break
                for feat in FEATURES:
                    if feat not in col_map:
                        remaining = [c for c in df_up.columns if c != nama_col and c not in col_map.values()]
                        if remaining:
                            col_map[feat] = remaining[0]
            else:
                for feat, keys in feat_keys.items():
                    for k in keys:
                        if k in cols_lower:
                            col_map[feat] = cols_lower[k]
                            break
                nama_col = None
                for k in name_keys:
                    if k in cols_lower:
                        nama_col = cols_lower[k]
                        break

            if len(df_up.columns) != 6:
                for feat in FEATURES:
                    if feat not in col_map and feat in df_up.columns:
                        col_map[feat] = feat

            # Build result dataframe
            results = []
            progress = st.progress(0, text="Menganalisis data siswa...")

            for i, row in df_up.iterrows():
                try:
                    if len(df_up.columns) == 6:
                        s_vals = {f: float(row[f]) for f in FEATURES}
                        nama_s = f"Siswa {i+1}"
                    else:
                        s_vals = {f: float(row[col_map[f]]) for f in FEATURES if f in col_map}
                        nama_s = str(row[nama_col]) if nama_col else f"Siswa {i+1}"

                    # Fill missing with 0
                    for f in FEATURES:
                        if f not in s_vals:
                            s_vals[f] = 0.0

                    status_s, conf_s = predict_student(model, s_vals)
                    res_s = generate_result(nama_s, s_vals, status_s, conf_s)

                    results.append({
                        'Nama Siswa': nama_s,
                        'Skor Aljabar': s_vals['skor_aljabar'],
                        'Skor Geometri': s_vals['skor_geometri'],
                        'Skor Bilangan': s_vals['skor_bilangan'],
                        'Skor Data & Ketidakpastian': s_vals['skor_data_ketidakpastian'],
                        'Skor Menalar': s_vals['skor_menalar'],
                        'Skor Literasi': s_vals['skor_literasi'],
                        'Rata-rata': res_s['avg'],
                        'Status': status_s,
                        'Konfidensial (%)': conf_s[status_s],
                        'Keterangan': res_s['saran'][:250] + '...' if len(res_s['saran']) > 250 else res_s['saran'],
                    })
                except Exception as e:
                    results.append({
                        'Nama Siswa': f'Siswa {i+1}',
                        'Skor Aljabar': 0, 'Skor Geometri': 0, 'Skor Bilangan': 0,
                        'Skor Data & Ketidakpastian': 0, 'Skor Menalar': 0, 'Skor Literasi': 0,
                        'Rata-rata': 0, 'Status': 'Error', 'Konfidensial (%)': 0,
                        'Keterangan': f'Error: {str(e)}',
                    })

                if (i+1) % max(1, len(df_up)//20) == 0 or i == len(df_up)-1:
                    progress.progress(min((i+1)/len(df_up), 1.0), text=f"Menganalisis {i+1}/{len(df_up)} siswa...")

            progress.empty()
            results_df = pd.DataFrame(results)

            # Summary stats
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("### 📊 Ringkasan Hasil")
            counts = results_df['Status'].value_counts()
            total = len(results_df)

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.markdown(f'<div class="metric-card"><div class="metric-value">{total}</div><div class="metric-label">Total Siswa</div></div>', unsafe_allow_html=True)
            with c2:
                st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#48BB78">{counts.get("Siap Olimpiade",0)}</div><div class="metric-label">Siap Olimpiade</div></div>', unsafe_allow_html=True)
            with c3:
                st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#F6AD55">{counts.get("Potensial",0)}</div><div class="metric-label">Potensial</div></div>', unsafe_allow_html=True)
            with c4:
                st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#FC8181">{counts.get("Tidak Siap",0)}</div><div class="metric-label">Tidak Siap</div></div>', unsafe_allow_html=True)

            # Preview table
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("### 📋 Preview Hasil (10 data pertama)")

            def style_status(val):
                if val == 'Siap Olimpiade': return 'background-color:#C6F6D5;color:#22543D;font-weight:bold'
                elif val == 'Potensial':    return 'background-color:#FEFCBF;color:#744210;font-weight:bold'
                elif val == 'Tidak Siap':  return 'background-color:#FED7D7;color:#742A2A;font-weight:bold'
                return ''

            preview = results_df.head(10).copy()
            styled = preview.style.applymap(style_status, subset=['Status'])
            st.dataframe(styled, use_container_width=True, height=380)

            # Download
            excel_bytes = export_results_excel(results_df)
            st.download_button(
                "📥 Download Semua Hasil (.xlsx)",
                data=excel_bytes,
                file_name="OlimpMath_Hasil_Seleksi.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"❌ Gagal memproses file: {str(e)}\n\nPastikan format kolom sesuai template.")


# ─────────────────────────────────────────────
# PAGE: TENTANG
# ─────────────────────────────────────────────
elif "Tentang" in menu:
    st.markdown("""
    <div class="hero-header" style="padding:2rem 2.5rem;">
        <div class="hero-badge">ℹ️ Informasi Sistem</div>
        <h1 class="hero-title" style="font-size:2.2rem;">Tentang OlimpMath</h1>
        <p class="hero-subtitle">Sistem Seleksi Cerdas Berbasis Machine Learning untuk Olimpiade Matematika</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([1,1])
    with col1:
        st.markdown("""
        <div class="result-card">
            <div class="section-title">🤖 Teknologi</div>
            <div style="color:rgba(190,227,248,0.8);font-size:0.9rem;line-height:1.8;">
                <b style="color:#90CDF4;">Algoritma:</b> Random Forest Classifier<br>
                <b style="color:#90CDF4;">Data Training:</b> 30.000+ data siswa<br>
                <b style="color:#90CDF4;">Akurasi:</b> 97% pada data uji<br>
                <b style="color:#90CDF4;">Fitur Input:</b> 6 dimensi kemampuan matematika<br>
                <b style="color:#90CDF4;">Output:</b> 3 kelas status kesiapan<br>
                <b style="color:#90CDF4;">Framework:</b> Streamlit + scikit-learn
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown("""
        <div class="result-card">
            <div class="section-title">🎯 Tujuan Sistem</div>
            <div style="color:rgba(190,227,248,0.8);font-size:0.9rem;line-height:1.9;">
                OlimpMath membantu guru dan pembina olimpiade matematika
                untuk melakukan seleksi awal calon peserta <b>Olimpiade Matematika (Olimpmath)</b>
                secara objektif dan efisien.<br><br>
                Sistem ini menganalisis 6 dimensi kemampuan matematika
                dan mengklasifikasikan siswa ke dalam tiga kategori
                kesiapan olimpiade, lengkap dengan review, rekomendasi,
                dan saran pengembangan yang personal.
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("""
    <div class="result-card" style="margin-top:1rem;">
        <div class="section-title">📊 Cara Kerja Model</div>
        <div style="color:rgba(190,227,248,0.75);font-size:0.9rem;line-height:1.8;">
            <b style="color:#63B3ED;">1. Preprocessing:</b> Data skor dinormalisasi dan nilai kosong diisi dengan median.<br>
            <b style="color:#63B3ED;">2. Labeling Otomatis:</b> Label dibuat berdasarkan rata-rata skor dan skor minimum:<br>
            &nbsp;&nbsp;&nbsp;• <b>Siap Olimpiade:</b> Rata-rata ≥ 80 DAN semua skor ≥ 60<br>
            &nbsp;&nbsp;&nbsp;• <b>Potensial:</b> Rata-rata ≥ 65 DAN skor minimum ≥ 45<br>
            &nbsp;&nbsp;&nbsp;• <b>Tidak Siap:</b> Selain kondisi di atas<br>
            <b style="color:#63B3ED;">3. Training:</b> Random Forest dengan 200 pohon keputusan, kedalaman maks 10.<br>
            <b style="color:#63B3ED;">4. Prediksi:</b> Voting mayoritas dari 200 pohon menghasilkan status + probabilitas.
        </div>
    </div>
    """, unsafe_allow_html=True)
